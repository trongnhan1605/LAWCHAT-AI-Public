from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


QUESTION_SHEETS = {
    "Questions_HNGD": "hon-nhan-va-gia-dinh",
    "Questions_DD": "dat-dai",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build benchmark cases from the legal QA tracking workbook.")
    parser.add_argument("--excel", type=Path, required=True)
    parser.add_argument("--output", type=Path, default=Path("benchmarks/generated_benchmark_cases.json"))
    parser.add_argument("--max-per-sheet", type=int, default=25)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook = openpyxl.load_workbook(args.excel, data_only=True)
    cases: list[dict[str, object]] = []
    for sheet_name, domain in QUESTION_SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        header_index = next((index for index, row in enumerate(rows) if "Prompt" in [str(cell).strip() for cell in row if cell is not None]), None)
        if header_index is None:
            continue
        headers = [str(cell or "").strip() for cell in rows[header_index]]
        prompt_index = headers.index("Prompt")
        expected_index = headers.index("Expected Output")
        legal_basis_index = headers.index("Legal Basic")
        category_index = headers.index("Category")
        stt_index = headers.index("STT")

        added = 0
        for row in rows[header_index + 1 :]:
            prompt = _cell_text(row[prompt_index] if prompt_index < len(row) else None)
            expected_output = _cell_text(row[expected_index] if expected_index < len(row) else None)
            if not prompt or not expected_output:
                continue
            category = _cell_text(row[category_index] if category_index < len(row) else None) or "Legal QA"
            stt = _cell_text(row[stt_index] if stt_index < len(row) else None) or str(added + 1)
            legal_basis = _cell_text(row[legal_basis_index] if legal_basis_index < len(row) else None)
            case_id = f"{'FAMILY' if domain == 'hon-nhan-va-gia-dinh' else 'LAND'}-XLSX-{int(float(stt)) if re.fullmatch(r'\\d+(\\.0)?', stt) else added + 1:03d}"
            cases.append(_build_case(case_id, domain, category, prompt, expected_output, legal_basis))
            added += 1
            if args.max_per_sheet and added >= args.max_per_sheet:
                break

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(cases, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"status": "generated", "output": str(args.output), "case_count": len(cases)}, ensure_ascii=False, indent=2))
    return 0


def _build_case(case_id: str, domain: str, category: str, prompt: str, expected_output: str, legal_basis: str | None) -> dict[str, object]:
    category_lower = category.lower()
    expected_complexity = "high" if any(term in category_lower for term in ["conflict", "temporal", "compliance"]) else "medium"
    expected_intent = "legal_qa"
    required_checks = ["answer_grounded"]
    must_escalate = False
    if "database coverage" in category_lower:
        expected_intent = "legal_search"
        required_checks = ["citation_present"]
    if "hierarchy" in category_lower or "conflict" in category_lower:
        expected_intent = "conflict_check"
        required_checks = ["citation_present", "conflict_checked", "legal_status_checked"]
    if "temporal" in category_lower:
        required_checks = ["citation_present", "legal_status_checked", "answer_grounded"]
    if "compliance" in category_lower:
        must_escalate = True
        required_checks = ["escalation_for_missing_evidence", "legal_status_checked"]

    return {
        "id": case_id,
        "domain": domain,
        "query": prompt,
        "expected_intent": expected_intent,
        "expected_complexity": expected_complexity,
        "expected_citations": _parse_legal_basis(legal_basis),
        "must_escalate": must_escalate,
        "required_checks": required_checks,
        "expected_answer_excerpt": expected_output[:700],
        "legal_basis_text": legal_basis or "",
        "source_category": category,
        "notes": "Generated from lawchat_benchmark_tracking.xlsx. Expected answer is benchmark guidance, not human-reviewed ground truth.",
    }


def _parse_legal_basis(legal_basis: str | None) -> list[dict[str, str | None]]:
    if not legal_basis:
        return []
    items = []
    for part in re.split(r"[;\n]+", legal_basis):
        text = part.strip(" -\t")
        if not text:
            continue
        items.append({"document_code": None, "article": text, "clause": None})
    return items


def _cell_text(value: object | None) -> str:
    return str(value or "").strip()


if __name__ == "__main__":
    raise SystemExit(main())
